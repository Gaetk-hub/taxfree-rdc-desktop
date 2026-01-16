"""
Reports Service for Customs Agents.
Provides data aggregation and Excel export functionality.
All data is strictly filtered by the agent's ID - each agent sees only their own data.
"""
import io
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Any, Optional
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from apps.customs.models import CustomsValidation, PointOfExit, ValidationDecision, AgentShift
from apps.taxfree.models import TaxFreeForm, TaxFreeFormStatus
from apps.refunds.models import Refund, RefundStatus


class CustomsReportsService:
    """
    Service for generating reports for customs agents.
    All queries are strictly filtered by agent_id - each agent sees only their own data.
    """
    
    def __init__(self, agent_id: str, point_of_exit_id: str, agent_name: str = None):
        """
        Initialize the service with the agent's ID.
        
        Args:
            agent_id: UUID of the agent (required - for filtering data)
            point_of_exit_id: UUID of the point of exit (frontier)
            agent_name: Name of the agent for display in reports
        """
        if not agent_id:
            raise ValueError("Agent ID is required")
        self.agent_id = agent_id
        self.point_of_exit_id = point_of_exit_id
        self.agent_name = agent_name
        self._validate_point_of_exit()
    
    def _validate_point_of_exit(self):
        """Validate that the point of exit exists."""
        if not PointOfExit.objects.filter(id=self.point_of_exit_id).exists():
            raise ValueError("Point of exit not found")
    
    # ==================== DASHBOARD STATISTICS ====================
    
    def get_dashboard_stats(self, date_from: datetime = None, date_to: datetime = None) -> Dict[str, Any]:
        """
        Get dashboard statistics for the point of exit.
        
        Args:
            date_from: Start date filter
            date_to: End date filter
            
        Returns:
            Dictionary with aggregated statistics
        """
        validations = self._get_validations_queryset(date_from, date_to)
        
        # Basic counts
        total_validations = validations.count()
        validated_count = validations.filter(decision=ValidationDecision.VALIDATED).count()
        refused_count = validations.filter(decision=ValidationDecision.REFUSED).count()
        
        # Amount statistics
        validated_forms = validations.filter(decision=ValidationDecision.VALIDATED)
        total_vat_validated = validated_forms.aggregate(
            total=Sum('form__vat_amount')
        )['total'] or Decimal('0')
        total_refund_amount = validated_forms.aggregate(
            total=Sum('form__refund_amount')
        )['total'] or Decimal('0')
        
        # Refund statistics
        refunds = self._get_refunds_queryset(date_from, date_to)
        refunds_paid = refunds.filter(status=RefundStatus.PAID)
        total_refunded = refunds_paid.aggregate(total=Sum('net_amount'))['total'] or Decimal('0')
        
        # Service gain statistics
        total_service_gain_cdf = refunds_paid.aggregate(total=Sum('service_gain_cdf'))['total'] or Decimal('0')
        
        # Count validated forms pending refund (VALIDATED status = waiting for refund)
        pending_refund_count = validated_forms.filter(form__status=TaxFreeFormStatus.VALIDATED).count()
        
        # Average processing time (validation to refund)
        avg_processing_time = None
        paid_refunds_with_time = refunds_paid.filter(
            paid_at__isnull=False,
            form__customs_validation__decided_at__isnull=False
        ).annotate(
            processing_time=F('paid_at') - F('form__customs_validation__decided_at')
        )
        if paid_refunds_with_time.exists():
            # Calculate average in seconds
            total_seconds = sum(
                r.processing_time.total_seconds() 
                for r in paid_refunds_with_time 
                if r.processing_time
            )
            avg_processing_time = total_seconds / paid_refunds_with_time.count() if paid_refunds_with_time.count() > 0 else 0
        
        return {
            'period': {
                'from': date_from.isoformat() if date_from else None,
                'to': date_to.isoformat() if date_to else None,
            },
            'validations': {
                'total': total_validations,
                'validated': validated_count,
                'refused': refused_count,
                'validation_rate': round(validated_count / total_validations * 100, 1) if total_validations > 0 else 0,
            },
            'amounts': {
                'total_vat_validated': float(total_vat_validated),
                'total_refund_amount': float(total_refund_amount),
                'total_refunded': float(total_refunded),
                'currency': 'CDF',
            },
            'refunds': {
                'total': refunds_paid.count() + pending_refund_count,
                'paid': refunds_paid.count(),
                'pending': pending_refund_count,
            },
            'service_gain': {
                'total_cdf': float(total_service_gain_cdf),
                'currency': 'CDF',
            },
            'performance': {
                'avg_processing_time_minutes': round(avg_processing_time / 60, 1) if avg_processing_time else None,
            }
        }
    
    def get_daily_stats(self, date_from: datetime = None, date_to: datetime = None) -> List[Dict[str, Any]]:
        """Get daily breakdown of validations and refunds."""
        validations = self._get_validations_queryset(date_from, date_to)
        
        # Get validation stats by day
        daily_validations = validations.annotate(
            date=TruncDate('decided_at')
        ).values('date').annotate(
            total=Count('id'),
            validated=Count('id', filter=Q(decision=ValidationDecision.VALIDATED)),
            refused=Count('id', filter=Q(decision=ValidationDecision.REFUSED)),
            total_amount=Sum('form__refund_amount', filter=Q(decision=ValidationDecision.VALIDATED))
        ).order_by('date')
        
        # Get refund stats by day (paid refunds)
        refunds = self._get_refunds_queryset(date_from, date_to)
        daily_refunds = refunds.filter(status=RefundStatus.PAID).annotate(
            date=TruncDate('paid_at')
        ).values('date').annotate(
            refunded_count=Count('id'),
            refunded_amount=Sum('net_amount')
        )
        
        # Create a dict for quick lookup of refund data
        refunds_by_date = {
            item['date']: {
                'refunded_count': item['refunded_count'],
                'refunded_amount': float(item['refunded_amount'] or 0)
            }
            for item in daily_refunds if item['date']
        }
        
        # Merge validation and refund data
        result = []
        for item in daily_validations:
            date = item['date']
            refund_data = refunds_by_date.get(date, {'refunded_count': 0, 'refunded_amount': 0})
            result.append({
                'date': date.isoformat() if date else None,
                'total': item['total'],
                'validated': item['validated'],
                'refused': item['refused'],
                'refunded': refund_data['refunded_count'],
                'total_amount': float(item['total_amount'] or 0),
                'refunded_amount': refund_data['refunded_amount'],
            })
        
        return result
    
    def get_agent_summary(self, date_from: datetime = None, date_to: datetime = None) -> Dict[str, Any]:
        """Get summary statistics for the current agent only."""
        validations = self._get_validations_queryset(date_from, date_to)
        refunds = self._get_refunds_queryset(date_from, date_to)
        
        total = validations.count()
        validated = validations.filter(decision=ValidationDecision.VALIDATED).count()
        refused = validations.filter(decision=ValidationDecision.REFUSED).count()
        total_amount = validations.filter(decision=ValidationDecision.VALIDATED).aggregate(
            total=Sum('form__refund_amount')
        )['total'] or Decimal('0')
        
        refunds_processed = refunds.filter(
            Q(initiated_by_id=self.agent_id) | Q(cash_collected_by_id=self.agent_id)
        ).count()
        refunds_paid = refunds.filter(status=RefundStatus.PAID).count()
        total_refunded = refunds.filter(status=RefundStatus.PAID).aggregate(
            total=Sum('net_amount')
        )['total'] or Decimal('0')
        
        return {
            'agent_id': str(self.agent_id),
            'agent_name': self.agent_name,
            'total_validations': total,
            'validated': validated,
            'refused': refused,
            'validation_rate': round(validated / total * 100, 1) if total > 0 else 0,
            'total_amount_validated': float(total_amount),
            'refunds_processed': refunds_processed,
            'refunds_paid': refunds_paid,
            'total_refunded': float(total_refunded),
        }
    
    # ==================== DETAILED DATA QUERIES ====================
    
    def get_validations_list(
        self, 
        date_from: datetime = None, 
        date_to: datetime = None,
        decision: str = None,
        limit: int = 1000
    ) -> List[Dict[str, Any]]:
        """Get detailed list of validations."""
        validations = self._get_validations_queryset(date_from, date_to)
        
        if decision:
            validations = validations.filter(decision=decision)
        
        validations = validations.select_related(
            'form', 'form__traveler', 'form__invoice', 'form__invoice__merchant',
            'agent', 'point_of_exit'
        )[:limit]
        
        return [
            {
                'id': str(v.id),
                'form_number': v.form.form_number,
                'decision': v.decision,
                'decision_display': v.get_decision_display(),
                'decided_at': v.decided_at.isoformat() if v.decided_at else None,
                'agent_name': v.agent.full_name,
                'agent_code': v.agent.agent_code,
                'traveler': {
                    'name': v.form.traveler.full_name if v.form.traveler else None,
                    'passport': f"***{v.form.traveler.passport_number_last4}" if v.form.traveler else None,
                    'nationality': v.form.traveler.nationality if v.form.traveler else None,
                },
                'merchant': {
                    'name': v.form.invoice.merchant.name if v.form.invoice and v.form.invoice.merchant else None,
                },
                'amounts': {
                    'eligible': float(v.form.eligible_amount),
                    'vat': float(v.form.vat_amount),
                    'refund': float(v.form.refund_amount),
                    'currency': v.form.currency,
                },
                'refusal_reason': v.refusal_reason if v.decision == ValidationDecision.REFUSED else None,
                'physical_control_done': v.physical_control_done,
                'is_offline': v.is_offline,
            }
            for v in validations
        ]
    
    def get_refunds_list(
        self,
        date_from: datetime = None,
        date_to: datetime = None,
        status: str = None,
        limit: int = 1000
    ) -> List[Dict[str, Any]]:
        """Get detailed list of refunds including failed/cancelled with reasons."""
        refunds = self._get_refunds_queryset(date_from, date_to)
        
        if status:
            refunds = refunds.filter(status=status)
        
        refunds = refunds.select_related(
            'form', 'form__traveler', 'form__invoice__merchant',
            'form__customs_validation',
            'initiated_by', 'cash_collected_by', 'cancelled_by'
        ).prefetch_related('attempts')[:limit]
        
        result = []
        for r in refunds:
            # Get last payment attempt error if failed
            last_error = None
            if r.status == 'FAILED':
                last_attempt = r.attempts.order_by('-started_at').first()
                if last_attempt:
                    last_error = {
                        'code': last_attempt.error_code,
                        'message': last_attempt.error_message,
                        'provider': last_attempt.provider,
                        'date': last_attempt.completed_at.isoformat() if last_attempt.completed_at else None,
                    }
            
            # Get validation refusal info if form was refused
            validation_refusal = None
            if hasattr(r.form, 'customs_validation') and r.form.customs_validation:
                cv = r.form.customs_validation
                if cv.decision == 'REFUSED':
                    validation_refusal = {
                        'reason': cv.refusal_reason,
                        'details': cv.refusal_details,
                        'agent': cv.agent.full_name if cv.agent else None,
                        'date': cv.decided_at.isoformat() if cv.decided_at else None,
                    }
            
            result.append({
                'id': str(r.id),
                'form_number': r.form.form_number,
                'status': r.status,
                'status_display': r.get_status_display(),
                'method': r.method,
                'method_display': r.get_method_display(),
                'traveler_name': r.form.traveler.full_name if r.form.traveler else None,
                'merchant_name': r.form.invoice.merchant.name if r.form.invoice and r.form.invoice.merchant else None,
                'amounts': {
                    'gross': float(r.gross_amount),
                    'fee': float(r.operator_fee),
                    'net': float(r.net_amount),
                    'currency': r.currency,
                },
                # Currency conversion info
                'payout': {
                    'currency': r.payout_currency or 'CDF',
                    'exchange_rate': float(r.exchange_rate_applied) if r.exchange_rate_applied else 1.0,
                    'amount': float(r.payout_amount) if r.payout_amount else float(r.net_amount),
                    'actual_amount': float(r.actual_payout_amount) if r.actual_payout_amount else None,
                },
                # Service gain info
                'service_gain': {
                    'amount': float(r.service_gain) if r.service_gain else 0.0,
                    'currency': r.payout_currency or 'CDF',
                    'amount_cdf': float(r.service_gain_cdf) if r.service_gain_cdf else 0.0,
                },
                'initiated_at': r.initiated_at.isoformat() if r.initiated_at else None,
                'initiated_by': r.initiated_by.full_name if r.initiated_by else None,
                'paid_at': r.paid_at.isoformat() if r.paid_at else None,
                'cash_collected_by': r.cash_collected_by.full_name if r.cash_collected_by else None,
                # Cancellation info
                'cancelled_at': r.cancelled_at.isoformat() if r.cancelled_at else None,
                'cancelled_by': r.cancelled_by.full_name if r.cancelled_by else None,
                'cancellation_reason': r.cancellation_reason or None,
                # Failure info
                'retry_count': r.retry_count,
                'last_error': last_error,
                # Validation refusal info (if form was refused at customs)
                'validation_refusal': validation_refusal,
            })
        
        return result
    
    # ==================== EXCEL EXPORT ====================
    
    def generate_validations_excel(
        self,
        date_from: datetime = None,
        date_to: datetime = None,
        decision: str = None
    ) -> io.BytesIO:
        """
        Generate Excel export of validations.
        
        Returns:
            BytesIO object containing the Excel file
        """
        validations = self.get_validations_list(date_from, date_to, decision, limit=10000)
        point_of_exit = PointOfExit.objects.get(id=self.point_of_exit_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Validations"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws.merge_cells('A1:M1')
        ws['A1'] = f"Mes Validations - {self.agent_name or 'Agent'}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Frontier row
        ws.merge_cells('A2:M2')
        ws['A2'] = f"Frontière: {point_of_exit.name} ({point_of_exit.code})"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Period row
        ws.merge_cells('A3:M3')
        period_text = "Période: "
        if date_from:
            period_text += f"Du {date_from.strftime('%d/%m/%Y')}"
        if date_to:
            period_text += f" au {date_to.strftime('%d/%m/%Y')}"
        if not date_from and not date_to:
            period_text += "Toutes les données"
        ws['A3'] = period_text
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # Generated date
        ws.merge_cells('A4:M4')
        ws['A4'] = f"Généré le: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            "N° Bordereau", "Date/Heure", "Décision", "Agent", "Code Agent",
            "Voyageur", "Passeport", "Nationalité", "Commerçant",
            "Montant Éligible", "TVA", "Remboursement", "Motif Refus"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, v in enumerate(validations, 7):
            ws.cell(row=row_idx, column=1, value=v['form_number']).border = thin_border
            ws.cell(row=row_idx, column=2, value=v['decided_at'][:19].replace('T', ' ') if v['decided_at'] else '').border = thin_border
            ws.cell(row=row_idx, column=3, value=v['decision_display']).border = thin_border
            ws.cell(row=row_idx, column=4, value=v['agent_name']).border = thin_border
            ws.cell(row=row_idx, column=5, value=v['agent_code'] or '').border = thin_border
            ws.cell(row=row_idx, column=6, value=v['traveler']['name'] or '').border = thin_border
            ws.cell(row=row_idx, column=7, value=v['traveler']['passport'] or '').border = thin_border
            ws.cell(row=row_idx, column=8, value=v['traveler']['nationality'] or '').border = thin_border
            ws.cell(row=row_idx, column=9, value=v['merchant']['name'] or '').border = thin_border
            ws.cell(row=row_idx, column=10, value=v['amounts']['eligible']).border = thin_border
            ws.cell(row=row_idx, column=11, value=v['amounts']['vat']).border = thin_border
            ws.cell(row=row_idx, column=12, value=v['amounts']['refund']).border = thin_border
            ws.cell(row=row_idx, column=13, value=v['refusal_reason'] or '').border = thin_border
            
            # Color coding for decision
            if v['decision'] == 'VALIDATED':
                ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif v['decision'] == 'REFUSED':
                ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Adjust column widths
        column_widths = [18, 18, 12, 20, 12, 25, 12, 10, 25, 15, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Summary section
        summary_row = len(validations) + 8
        ws.cell(row=summary_row, column=1, value="RÉSUMÉ").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value="Total validations:")
        ws.cell(row=summary_row + 1, column=2, value=len(validations))
        ws.cell(row=summary_row + 2, column=1, value="Validés:")
        ws.cell(row=summary_row + 2, column=2, value=sum(1 for v in validations if v['decision'] == 'VALIDATED'))
        ws.cell(row=summary_row + 3, column=1, value="Refusés:")
        ws.cell(row=summary_row + 3, column=2, value=sum(1 for v in validations if v['decision'] == 'REFUSED'))
        ws.cell(row=summary_row + 4, column=1, value="Total TVA validée:")
        ws.cell(row=summary_row + 4, column=2, value=sum(v['amounts']['vat'] for v in validations if v['decision'] == 'VALIDATED'))
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_refunds_excel(
        self,
        date_from: datetime = None,
        date_to: datetime = None,
        status: str = None
    ) -> io.BytesIO:
        """Generate Excel export of refunds."""
        refunds = self.get_refunds_list(date_from, date_to, status, limit=10000)
        point_of_exit = PointOfExit.objects.get(id=self.point_of_exit_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Remboursements"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:K1')
        ws['A1'] = f"Mes Remboursements - {self.agent_name or 'Agent'}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Frontier row
        ws.merge_cells('A2:K2')
        ws['A2'] = f"Frontière: {point_of_exit.name} ({point_of_exit.code})"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Period
        ws.merge_cells('A3:K3')
        period_text = "Période: "
        if date_from:
            period_text += f"Du {date_from.strftime('%d/%m/%Y')}"
        if date_to:
            period_text += f" au {date_to.strftime('%d/%m/%Y')}"
        if not date_from and not date_to:
            period_text += "Toutes les données"
        ws['A3'] = period_text
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # Generated date
        ws.merge_cells('A4:K4')
        ws['A4'] = f"Généré le: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'].alignment = Alignment(horizontal="center")
        
        # Headers - including currency conversion, service gain and cancellation/failure columns
        headers = [
            "N° Bordereau", "Statut", "Méthode", "Voyageur", "Commerçant",
            "Montant Brut (CDF)", "Frais (CDF)", "Montant Net (CDF)", 
            "Devise Paiement", "Taux Change", "Montant Prévu", "Montant Donné", "Gain Service", "Gain Service (CDF)",
            "Initié le", "Payé le", "Agent", "Annulé le", "Motif Annulation/Échec"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, r in enumerate(refunds, 7):
            ws.cell(row=row_idx, column=1, value=r['form_number']).border = thin_border
            ws.cell(row=row_idx, column=2, value=r['status_display']).border = thin_border
            ws.cell(row=row_idx, column=3, value=r['method_display']).border = thin_border
            ws.cell(row=row_idx, column=4, value=r['traveler_name'] or '').border = thin_border
            ws.cell(row=row_idx, column=5, value=r['merchant_name'] or '').border = thin_border
            ws.cell(row=row_idx, column=6, value=r['amounts']['gross']).border = thin_border
            ws.cell(row=row_idx, column=7, value=r['amounts']['fee']).border = thin_border
            ws.cell(row=row_idx, column=8, value=r['amounts']['net']).border = thin_border
            # Currency conversion columns
            payout = r.get('payout', {})
            service_gain = r.get('service_gain', {})
            ws.cell(row=row_idx, column=9, value=payout.get('currency', 'CDF')).border = thin_border
            ws.cell(row=row_idx, column=10, value=payout.get('exchange_rate', 1.0)).border = thin_border
            ws.cell(row=row_idx, column=11, value=payout.get('amount', r['amounts']['net'])).border = thin_border
            ws.cell(row=row_idx, column=12, value=payout.get('actual_amount') or payout.get('amount', r['amounts']['net'])).border = thin_border
            ws.cell(row=row_idx, column=13, value=service_gain.get('amount', 0)).border = thin_border
            ws.cell(row=row_idx, column=14, value=service_gain.get('amount_cdf', 0)).border = thin_border
            ws.cell(row=row_idx, column=15, value=r['initiated_at'][:10] if r['initiated_at'] else '').border = thin_border
            ws.cell(row=row_idx, column=16, value=r['paid_at'][:10] if r['paid_at'] else '').border = thin_border
            ws.cell(row=row_idx, column=17, value=r['initiated_by'] or r['cash_collected_by'] or '').border = thin_border
            
            # Cancellation/Failure date
            cancel_date = ''
            if r.get('cancelled_at'):
                cancel_date = r['cancelled_at'][:10]
            elif r.get('last_error') and r['last_error'].get('date'):
                cancel_date = r['last_error']['date'][:10]
            ws.cell(row=row_idx, column=18, value=cancel_date).border = thin_border
            
            # Cancellation/Failure reason
            reason = ''
            if r.get('cancellation_reason'):
                reason = r['cancellation_reason']
            elif r.get('last_error'):
                error = r['last_error']
                reason = f"{error.get('message', '')} ({error.get('code', '')})"
            ws.cell(row=row_idx, column=19, value=reason).border = thin_border
            
            # Color coding for status
            if r['status'] == 'PAID':
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif r['status'] in ['PENDING', 'INITIATED']:
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif r['status'] in ['FAILED', 'CANCELLED']:
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Adjust column widths
        column_widths = [18, 12, 15, 25, 25, 15, 12, 15, 10, 12, 15, 15, 12, 15, 12, 12, 20, 12, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Summary
        summary_row = len(refunds) + 9
        ws.cell(row=summary_row, column=1, value="RÉSUMÉ").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value="Total remboursements:")
        ws.cell(row=summary_row + 1, column=2, value=len(refunds))
        ws.cell(row=summary_row + 2, column=1, value="Payés:")
        ws.cell(row=summary_row + 2, column=2, value=sum(1 for r in refunds if r['status'] == 'PAID'))
        ws.cell(row=summary_row + 3, column=1, value="En attente:")
        ws.cell(row=summary_row + 3, column=2, value=sum(1 for r in refunds if r['status'] in ['PENDING', 'INITIATED']))
        ws.cell(row=summary_row + 4, column=1, value="Échoués:")
        ws.cell(row=summary_row + 4, column=2, value=sum(1 for r in refunds if r['status'] == 'FAILED'))
        ws.cell(row=summary_row + 5, column=1, value="Annulés:")
        ws.cell(row=summary_row + 5, column=2, value=sum(1 for r in refunds if r['status'] == 'CANCELLED'))
        ws.cell(row=summary_row + 6, column=1, value="Total remboursé (CDF):")
        ws.cell(row=summary_row + 6, column=2, value=sum(r['amounts']['net'] for r in refunds if r['status'] == 'PAID'))
        ws.cell(row=summary_row + 7, column=1, value="Total gain de service (CDF):").font = Font(bold=True, color="0000FF")
        ws.cell(row=summary_row + 7, column=2, value=sum(r.get('service_gain', {}).get('amount_cdf', 0) for r in refunds if r['status'] == 'PAID')).font = Font(bold=True, color="0000FF")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_summary_excel(
        self,
        date_from: datetime = None,
        date_to: datetime = None
    ) -> io.BytesIO:
        """Generate comprehensive summary Excel report for the agent."""
        point_of_exit = PointOfExit.objects.get(id=self.point_of_exit_id)
        stats = self.get_dashboard_stats(date_from, date_to)
        daily_stats = self.get_daily_stats(date_from, date_to)
        agent_summary = self.get_agent_summary(date_from, date_to)
        
        wb = Workbook()
        
        # ===== Sheet 1: Summary =====
        ws1 = wb.active
        ws1.title = "Résumé"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        
        ws1.merge_cells('A1:D1')
        ws1['A1'] = f"Rapport Personnel - {self.agent_name or 'Agent'}"
        ws1['A1'].font = Font(bold=True, size=16)
        
        ws1.merge_cells('A2:D2')
        ws1['A2'] = f"Frontière: {point_of_exit.name} ({point_of_exit.code})"
        
        ws1.merge_cells('A3:D3')
        period_text = "Période: "
        if date_from:
            period_text += f"Du {date_from.strftime('%d/%m/%Y')}"
        if date_to:
            period_text += f" au {date_to.strftime('%d/%m/%Y')}"
        if not date_from and not date_to:
            period_text += "Toutes les données"
        ws1['A3'] = period_text
        
        ws1['A5'] = "MES STATISTIQUES DE VALIDATION"
        ws1['A5'].font = Font(bold=True, size=12)
        
        ws1['A7'] = "Total validations effectuées"
        ws1['B7'] = stats['validations']['total']
        ws1['A8'] = "Bordereaux validés"
        ws1['B8'] = stats['validations']['validated']
        ws1['A9'] = "Bordereaux refusés"
        ws1['B9'] = stats['validations']['refused']
        ws1['A10'] = "Mon taux de validation"
        ws1['B10'] = f"{stats['validations']['validation_rate']}%"
        
        ws1['A12'] = "MES MONTANTS TRAITÉS"
        ws1['A12'].font = Font(bold=True, size=12)
        
        ws1['A14'] = "TVA totale validée"
        ws1['B14'] = f"{stats['amounts']['total_vat_validated']:,.0f} CDF"
        ws1['A15'] = "Montant à rembourser"
        ws1['B15'] = f"{stats['amounts']['total_refund_amount']:,.0f} CDF"
        ws1['A16'] = "Montant remboursé"
        ws1['B16'] = f"{stats['amounts']['total_refunded']:,.0f} CDF"
        
        ws1['A18'] = "MES REMBOURSEMENTS"
        ws1['A18'].font = Font(bold=True, size=12)
        
        ws1['A20'] = "Total remboursements traités"
        ws1['B20'] = stats['refunds']['total']
        ws1['A21'] = "Payés"
        ws1['B21'] = stats['refunds']['paid']
        ws1['A22'] = "En attente"
        ws1['B22'] = stats['refunds']['pending']
        
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20
        
        # ===== Sheet 2: Daily Stats =====
        ws2 = wb.create_sheet("Statistiques Journalières")
        
        headers = ["Date", "Total", "Validés", "Refusés", "Remboursés", "Montant Validé (CDF)", "Montant Remboursé (CDF)"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, day in enumerate(daily_stats, 2):
            ws2.cell(row=row_idx, column=1, value=day['date'])
            ws2.cell(row=row_idx, column=2, value=day['total'])
            ws2.cell(row=row_idx, column=3, value=day['validated'])
            ws2.cell(row=row_idx, column=4, value=day['refused'])
            ws2.cell(row=row_idx, column=5, value=day.get('refunded', 0))
            ws2.cell(row=row_idx, column=6, value=day['total_amount'])
            ws2.cell(row=row_idx, column=7, value=day.get('refunded_amount', 0))
        
        for i, width in enumerate([12, 10, 10, 10, 12, 18, 20], 1):
            ws2.column_dimensions[get_column_letter(i)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ==================== PRIVATE HELPER METHODS ====================
    
    def _get_validations_queryset(self, date_from: datetime = None, date_to: datetime = None):
        """Get base queryset for validations filtered by agent ID."""
        # Filter strictly by the agent who performed the validation
        queryset = CustomsValidation.objects.filter(
            agent_id=self.agent_id
        )
        
        if date_from:
            queryset = queryset.filter(decided_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(decided_at__lte=date_to)
        
        return queryset
    
    def _get_refunds_queryset(self, date_from: datetime = None, date_to: datetime = None):
        """Get base queryset for refunds filtered by agent ID."""
        # Refunds are linked to forms that were validated by this agent
        # OR refunds that were initiated/collected by this agent
        queryset = Refund.objects.filter(
            Q(form__customs_validation__agent_id=self.agent_id) |
            Q(initiated_by_id=self.agent_id) |
            Q(cash_collected_by_id=self.agent_id)
        ).distinct()
        
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        return queryset

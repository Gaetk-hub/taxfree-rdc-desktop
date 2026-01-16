"""
Views for audit app - Advanced audit trail with comprehensive filtering.
"""
from rest_framework import viewsets, views, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncDate, TruncHour
from django.utils import timezone
from datetime import timedelta
import csv
import io

from apps.accounts.permissions import IsAdmin, IsAdminOrAuditor
from .models import AuditLog
from .serializers import AuditLogSerializer, AuditLogDetailSerializer


# Action categories for better filtering
ACTION_CATEGORIES = {
    'AUTHENTICATION': ['LOGIN', 'LOGOUT', 'PASSWORD_RESET', 'OTP_SENT', 'OTP_VERIFIED'],
    'USER_MANAGEMENT': ['USER_CREATED', 'USER_UPDATED', 'USER_DEACTIVATED', 'UPDATE_CUSTOMS_AGENT', 
                        'DEACTIVATE_CUSTOMS_AGENT', 'ACTIVATE_AGENT_ACCOUNT'],
    'MERCHANT': ['MERCHANT_CREATED', 'MERCHANT_UPDATED', 'MERCHANT_APPROVED', 'MERCHANT_REJECTED',
                 'OUTLET_CREATED', 'OUTLET_UPDATED', 'OUTLET_ACTIVATED', 'OUTLET_DEACTIVATED',
                 'MERCHANT_USER_INVITED', 'MERCHANT_INVITATION_ACCEPTED'],
    'TAXFREE_FORMS': ['TAXFREE_FORM_CREATED', 'TAXFREE_FORM_CANCELLED', 'TAXFREE_FORM_EMAILED',
                      'TAXFREE_FORM_PDF_DOWNLOADED', 'SCAN_FORM', 'LOOKUP_FORM', 'SEARCH_FORMS'],
    'CUSTOMS': ['CUSTOMS_VALIDATED', 'CUSTOMS_REFUSED', 'START_SHIFT', 'END_SHIFT', 
                'PAUSE_SHIFT', 'RESUME_SHIFT', 'OFFLINE_SYNC', 'CREATE_AGENT_INVITATION'],
    'REFUNDS': ['REFUND_INITIATED', 'REFUND_COMPLETED', 'REFUND_FAILED', 'REFUND_CANCELLED',
                'CASH_COLLECTED', 'RECEIPT_SENT'],
    'ADMIN': ['STATUS_OVERRIDE', 'RULE_CREATED', 'RULE_UPDATED', 'CATEGORY_CREATED', 
              'PERMISSION_UPDATED', 'SYSTEM_CONFIG_UPDATED'],
    'INVOICES': ['INVOICE_CREATED', 'INVOICE_CANCELLED'],
}

# Entity display names
ENTITY_DISPLAY = {
    'TaxFreeForm': 'Bordereau Tax Free',
    'User': 'Utilisateur',
    'Merchant': 'Commerçant',
    'Outlet': 'Point de vente',
    'SaleInvoice': 'Facture',
    'Refund': 'Remboursement',
    'CustomsValidation': 'Validation douanière',
    'AgentShift': 'Vacation agent',
    'CustomsAgentInvitation': 'Invitation agent',
    'MerchantUserInvitation': 'Invitation commerçant',
    'PointOfExit': 'Point de sortie',
    'TaxRule': 'Règle TVA',
    'ProductCategory': 'Catégorie produit',
}

# Role display names
ROLE_DISPLAY = {
    'ADMIN': 'Super Admin',
    'AUDITOR': 'Auditeur',
    'OPERATOR': 'Opérateur',
    'CUSTOMS_AGENT': 'Agent douanier',
    'MERCHANT': 'Commerçant',
    'MERCHANT_EMPLOYEE': 'Employé commerçant',
    'CLIENT': 'Client',
    '': 'Système',
}


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Advanced ViewSet for audit logs with comprehensive filtering.
    Supports filtering by action, entity, actor, role, date range, IP, and more.
    """
    
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsAdminOrAuditor]
    ordering_fields = ['timestamp']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AuditLogDetailSerializer
        return AuditLogSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        
        # Text search (email, entity_id, action)
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(actor_email__icontains=search) |
                Q(entity_id__icontains=search) |
                Q(action__icontains=search) |
                Q(metadata__icontains=search)
            )
        
        # Filter by action(s)
        action = params.get('action')
        if action:
            actions = action.split(',')
            queryset = queryset.filter(action__in=actions)
        
        # Filter by action category
        category = params.get('category')
        if category and category in ACTION_CATEGORIES:
            queryset = queryset.filter(action__in=ACTION_CATEGORIES[category])
        
        # Filter by entity type(s)
        entity = params.get('entity')
        if entity:
            entities = entity.split(',')
            queryset = queryset.filter(entity__in=entities)
        
        # Filter by actor ID
        actor_id = params.get('actor_id')
        if actor_id:
            queryset = queryset.filter(actor_id=actor_id)
        
        # Filter by actor email
        actor_email = params.get('actor_email')
        if actor_email:
            queryset = queryset.filter(actor_email__icontains=actor_email)
        
        # Filter by role(s)
        role = params.get('role')
        if role:
            roles = role.split(',')
            queryset = queryset.filter(actor_role__in=roles)
        
        # Filter by IP address
        ip = params.get('ip')
        if ip:
            queryset = queryset.filter(actor_ip__icontains=ip)
        
        # Date range filter
        start_date = params.get('start_date')
        end_date = params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(timestamp__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(timestamp__date__lte=end_date)
        
        # Quick date filters
        period = params.get('period')
        if period:
            now = timezone.now()
            if period == 'today':
                queryset = queryset.filter(timestamp__date=now.date())
            elif period == 'yesterday':
                queryset = queryset.filter(timestamp__date=now.date() - timedelta(days=1))
            elif period == 'week':
                queryset = queryset.filter(timestamp__gte=now - timedelta(days=7))
            elif period == 'month':
                queryset = queryset.filter(timestamp__gte=now - timedelta(days=30))
        
        # Filter by entity_id (specific entity)
        entity_id = params.get('entity_id')
        if entity_id:
            queryset = queryset.filter(entity_id__icontains=entity_id)
        
        return queryset.order_by('-timestamp')

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get audit statistics for dashboard."""
        params = request.query_params
        days = int(params.get('days', 30))
        since = timezone.now() - timedelta(days=days)
        
        queryset = AuditLog.objects.filter(timestamp__gte=since)
        
        # Total count
        total = queryset.count()
        
        # By action
        by_action = dict(queryset.values('action').annotate(
            count=Count('id')
        ).order_by('-count').values_list('action', 'count')[:15])
        
        # By entity
        by_entity = dict(queryset.values('entity').annotate(
            count=Count('id')
        ).order_by('-count').values_list('entity', 'count'))
        
        # By role
        by_role = dict(queryset.values('actor_role').annotate(
            count=Count('id')
        ).order_by('-count').values_list('actor_role', 'count'))
        
        # By day (for chart)
        by_day = list(queryset.annotate(
            date=TruncDate('timestamp')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date').values('date', 'count'))
        
        # Top actors
        top_actors = list(queryset.exclude(actor_email='').values(
            'actor_id', 'actor_email', 'actor_role'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:10])
        
        # Recent critical actions
        critical_actions = ['STATUS_OVERRIDE', 'USER_DEACTIVATED', 'MERCHANT_REJECTED', 
                          'CUSTOMS_REFUSED', 'REFUND_CANCELLED']
        recent_critical = list(queryset.filter(
            action__in=critical_actions
        ).order_by('-timestamp')[:5].values(
            'id', 'action', 'entity', 'entity_id', 'actor_email', 'timestamp'
        ))
        
        return Response({
            'total': total,
            'period_days': days,
            'by_action': by_action,
            'by_entity': by_entity,
            'by_role': by_role,
            'by_day': by_day,
            'top_actors': top_actors,
            'recent_critical': recent_critical,
        })

    @action(detail=False, methods=['get'])
    def filters(self, request):
        """Get available filter options."""
        # Get unique values for filters
        actions = list(AuditLog.objects.values_list('action', flat=True).distinct().order_by('action'))
        entities = list(AuditLog.objects.values_list('entity', flat=True).distinct().order_by('entity'))
        roles = list(AuditLog.objects.exclude(actor_role='').values_list('actor_role', flat=True).distinct())
        
        return Response({
            'actions': actions,
            'entities': entities,
            'roles': roles,
            'action_categories': {k: v for k, v in ACTION_CATEGORIES.items()},
            'entity_display': ENTITY_DISPLAY,
            'role_display': ROLE_DISPLAY,
        })

    @action(detail=False, methods=['get'])
    def entity_history(self, request):
        """Get complete history for a specific entity."""
        entity = request.query_params.get('entity')
        entity_id = request.query_params.get('entity_id')
        
        if not entity or not entity_id:
            return Response(
                {'detail': 'entity and entity_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        logs = AuditLog.objects.filter(
            entity=entity,
            entity_id__icontains=entity_id
        ).order_by('-timestamp')[:100]
        
        serializer = AuditLogDetailSerializer(logs, many=True)
        return Response({
            'entity': entity,
            'entity_id': entity_id,
            'history': serializer.data,
            'count': logs.count(),
        })

    @action(detail=False, methods=['get'])
    def actor_history(self, request):
        """Get complete history for a specific actor."""
        actor_id = request.query_params.get('actor_id')
        actor_email = request.query_params.get('actor_email')
        
        if not actor_id and not actor_email:
            return Response(
                {'detail': 'actor_id or actor_email is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = AuditLog.objects.all()
        if actor_id:
            queryset = queryset.filter(actor_id=actor_id)
        elif actor_email:
            queryset = queryset.filter(actor_email__icontains=actor_email)
        
        logs = queryset.order_by('-timestamp')[:100]
        
        serializer = AuditLogDetailSerializer(logs, many=True)
        
        # Get actor summary
        summary = queryset.values('action').annotate(count=Count('id')).order_by('-count')
        
        return Response({
            'actor_id': actor_id,
            'actor_email': actor_email,
            'history': serializer.data,
            'count': logs.count(),
            'action_summary': list(summary),
        })

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export audit logs to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return self._export_csv(request)
        
        queryset = self.get_queryset()[:5000]  # Limit export
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Journal d\'Audit'
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        
        # Headers
        headers = ['Date/Heure', 'Action', 'Catégorie', 'Entité', 'ID Entité', 
                   'Utilisateur', 'Rôle', 'Adresse IP', 'Détails']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data
        for row, log in enumerate(queryset, 2):
            # Find category
            category = 'Autre'
            for cat, actions in ACTION_CATEGORIES.items():
                if log.action in actions:
                    category = cat
                    break
            
            ws.cell(row=row, column=1, value=log.timestamp.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row, column=2, value=log.action)
            ws.cell(row=row, column=3, value=category)
            ws.cell(row=row, column=4, value=ENTITY_DISPLAY.get(log.entity, log.entity))
            ws.cell(row=row, column=5, value=log.entity_id)
            ws.cell(row=row, column=6, value=log.actor_email or 'Système')
            ws.cell(row=row, column=7, value=ROLE_DISPLAY.get(log.actor_role, log.actor_role))
            ws.cell(row=row, column=8, value=log.actor_ip or '-')
            ws.cell(row=row, column=9, value=str(log.metadata)[:500] if log.metadata else '-')
        
        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        ws.freeze_panes = 'A2'
        
        # Response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"audit_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _export_csv(self, request):
        """Fallback CSV export."""
        queryset = self.get_queryset()[:5000]
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date/Heure', 'Action', 'Entité', 'ID Entité', 
                        'Utilisateur', 'Rôle', 'IP', 'Détails'])
        
        for log in queryset:
            writer.writerow([
                log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                log.action,
                log.entity,
                log.entity_id,
                log.actor_email or 'Système',
                log.actor_role,
                log.actor_ip or '-',
                str(log.metadata)[:200] if log.metadata else '-',
            ])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="audit_export.csv"'
        return response

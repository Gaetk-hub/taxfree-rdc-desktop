"""
Admin views for TaxFree app - Full traceability for Super Admin.
"""
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count, Sum
from django.utils import timezone

from apps.accounts.permissions import IsAdmin, IsAdminOrAuditor, PermissionService
from apps.refunds.models import Refund, RefundStatus
from .models import TaxFreeForm, TaxFreeFormStatus
from .admin_serializers import (
    AdminTaxFreeFormListSerializer,
    AdminTaxFreeFormDetailSerializer,
    AdminRefundListSerializer,
    AdminRefundDetailSerializer,
)


class AdminTaxFreeFormViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Admin ViewSet for TaxFree Forms with full traceability.
    Provides advanced search, filtering, and detailed views.
    """
    
    queryset = TaxFreeForm.objects.all()
    permission_classes = [IsAuthenticated, IsAdminOrAuditor]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'requires_control', 'currency']
    search_fields = [
        'form_number', 
        'traveler__first_name', 
        'traveler__last_name',
        'traveler__passport_number_last4',
        'invoice__merchant__name',
        'invoice__outlet__name',
        'invoice__invoice_number',
    ]
    ordering_fields = [
        'created_at', 'issued_at', 'validated_at', 'expires_at',
        'refund_amount', 'eligible_amount', 'risk_score'
    ]
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AdminTaxFreeFormDetailSerializer
        return AdminTaxFreeFormListSerializer

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'traveler',
            'invoice',
            'invoice__merchant',
            'invoice__outlet',
            'created_by',
            'cancelled_by',
        ).prefetch_related(
            'invoice__items',
        )
        
        # Advanced filters from query params
        params = self.request.query_params
        
        # Date range filters
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        # Merchant filter
        merchant_id = params.get('merchant')
        if merchant_id:
            queryset = queryset.filter(invoice__merchant_id=merchant_id)
        
        # Outlet filter
        outlet_id = params.get('outlet')
        if outlet_id:
            queryset = queryset.filter(invoice__outlet_id=outlet_id)
        
        # Validation status filter
        validation_status = params.get('validation_status')
        if validation_status == 'validated':
            queryset = queryset.filter(customs_validation__decision='VALIDATED')
        elif validation_status == 'refused':
            queryset = queryset.filter(customs_validation__decision='REFUSED')
        elif validation_status == 'pending':
            queryset = queryset.filter(
                status__in=[TaxFreeFormStatus.ISSUED, TaxFreeFormStatus.VALIDATION_PENDING]
            ).exclude(customs_validation__isnull=False)
        
        # Refund status filter
        refund_status = params.get('refund_status')
        if refund_status == 'paid':
            queryset = queryset.filter(refund__status=RefundStatus.PAID)
        elif refund_status == 'pending':
            queryset = queryset.filter(refund__status=RefundStatus.PENDING)
        elif refund_status == 'failed':
            queryset = queryset.filter(refund__status=RefundStatus.FAILED)
        elif refund_status == 'none':
            queryset = queryset.filter(refund__isnull=True)
        
        # Risk level filter
        risk_level = params.get('risk_level')
        if risk_level == 'high':
            queryset = queryset.filter(risk_score__gte=70)
        elif risk_level == 'medium':
            queryset = queryset.filter(risk_score__gte=40, risk_score__lt=70)
        elif risk_level == 'low':
            queryset = queryset.filter(risk_score__lt=40)
        
        # Expired filter
        is_expired = params.get('is_expired')
        if is_expired == 'true':
            queryset = queryset.filter(expires_at__lt=timezone.now())
        elif is_expired == 'false':
            queryset = queryset.filter(expires_at__gte=timezone.now())
        
        # Amount range filters
        min_amount = params.get('min_amount')
        max_amount = params.get('max_amount')
        if min_amount:
            queryset = queryset.filter(refund_amount__gte=min_amount)
        if max_amount:
            queryset = queryset.filter(refund_amount__lte=max_amount)
        
        return queryset

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get statistics for the forms dashboard."""
        queryset = self.get_queryset()
        
        # Apply same filters as list
        params = request.query_params
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        # Status counts
        status_counts = {}
        for status_choice in TaxFreeFormStatus.choices:
            status_counts[status_choice[0]] = queryset.filter(status=status_choice[0]).count()
        
        # Totals
        totals = queryset.aggregate(
            total_count=Count('id'),
            total_eligible=Sum('eligible_amount'),
            total_vat=Sum('vat_amount'),
            total_refund=Sum('refund_amount'),
        )
        
        # Validation stats
        validated_count = queryset.filter(customs_validation__decision='VALIDATED').count()
        refused_count = queryset.filter(customs_validation__decision='REFUSED').count()
        pending_validation = queryset.filter(
            status__in=[TaxFreeFormStatus.ISSUED, TaxFreeFormStatus.VALIDATION_PENDING]
        ).exclude(customs_validation__isnull=False).count()
        
        # Refund stats
        refund_paid = queryset.filter(refund__status=RefundStatus.PAID).count()
        refund_pending = queryset.filter(refund__status=RefundStatus.PENDING).count()
        refund_failed = queryset.filter(refund__status=RefundStatus.FAILED).count()
        
        # Risk stats
        high_risk = queryset.filter(risk_score__gte=70).count()
        control_required = queryset.filter(requires_control=True).count()
        
        # Expired
        expired = queryset.filter(expires_at__lt=timezone.now()).count()
        
        return Response({
            'status_counts': status_counts,
            'totals': {
                'count': totals['total_count'] or 0,
                'eligible_amount': float(totals['total_eligible'] or 0),
                'vat_amount': float(totals['total_vat'] or 0),
                'refund_amount': float(totals['total_refund'] or 0),
            },
            'validation': {
                'validated': validated_count,
                'refused': refused_count,
                'pending': pending_validation,
            },
            'refund': {
                'paid': refund_paid,
                'pending': refund_pending,
                'failed': refund_failed,
            },
            'risk': {
                'high_risk': high_risk,
                'control_required': control_required,
            },
            'expired': expired,
        })

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export forms to comprehensive Excel file with all details."""
        # Check granular permission for EXPORT action
        if request.user.role in ['ADMIN', 'AUDITOR'] and not getattr(request.user, 'is_super_admin', False):
            if not PermissionService.has_module_permission(request.user, 'FORMS', 'EXPORT'):
                return Response(
                    {'error': 'Vous n\'avez pas la permission d\'exporter les bordereaux.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        import io
        from django.http import HttpResponse
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return self._export_csv(request)
        
        queryset = self.get_queryset().select_related(
            'traveler',
            'invoice__merchant',
            'invoice__outlet',
            'invoice__created_by',
            'created_by',
            'cancelled_by',
        ).prefetch_related(
            'invoice__items',
        )[:2000]  # Limit export
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== SHEET 1: BORDEREAUX ====================
        ws_forms = wb.active
        ws_forms.title = 'Bordereaux'
        
        forms_headers = [
            'N° Bordereau', 'Statut', 'Devise', 'Montant Éligible', 'TVA', 'Frais Opérateur',
            'Montant Remboursement', 'Score Risque', 'Contrôle Requis', 'Indicateurs Risque',
            'Expiré', 'Créé le', 'Émis le', 'Expire le', 'Validé le', 'Annulé le',
            'Raison Annulation', 'Créé par', 'Annulé par',
            # Voyageur
            'Voyageur Nom', 'Voyageur Prénom', 'Passeport', 'Pays Passeport',
            'Date Naissance', 'Nationalité', 'Pays Résidence', 'Email Voyageur', 'Tél Voyageur',
            'Méthode Remb. Préférée',
            # Facture
            'N° Facture', 'Date Facture', 'Sous-total Facture', 'TVA Facture', 'Total Facture',
            'Mode Paiement', 'Réf. Paiement',
            # Commerçant
            'Commerçant', 'Nom Commercial', 'RCCM', 'NIF', 'Ville Commerçant', 'Province Commerçant',
            'Contact Commerçant', 'Email Commerçant', 'Tél Commerçant',
            # Point de vente
            'Point de Vente', 'Code PDV', 'Adresse PDV', 'Ville PDV',
            # Validation douanière
            'Décision Douane', 'Motif Refus', 'Détails Refus', 'Contrôle Physique',
            'Notes Contrôle', 'Agent Douane', 'Code Agent', 'Email Agent',
            'Point de Sortie', 'Code Frontière', 'Type Frontière', 'Ville Frontière',
            'Validation Hors-ligne', 'Date Validation',
            # Remboursement
            'Statut Remboursement', 'Méthode Remboursement', 'Montant Brut', 'Frais',
            'Montant Net', 'Devise Paiement', 'Taux Change', 'Montant Attendu',
            'Montant Réellement Payé', 'Gain Service', 'Gain Service (CDF)',
            'Initié le', 'Initié par', 'Payé le',
            'Espèces Collectées', 'Collecté le', 'Collecté par', 'Tentatives'
        ]
        
        # Write headers
        for col, header in enumerate(forms_headers, 1):
            cell = ws_forms.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Freeze header row
        ws_forms.freeze_panes = 'A2'
        
        # Write data
        for row_idx, form in enumerate(queryset, 2):
            # Get related objects safely
            traveler = form.traveler
            invoice = form.invoice
            merchant = invoice.merchant if invoice else None
            outlet = invoice.outlet if invoice else None
            
            # Customs validation
            try:
                validation = form.customs_validation
                val_decision = validation.get_decision_display()
                val_refusal = validation.get_refusal_reason_display() if validation.refusal_reason else ''
                val_refusal_details = validation.refusal_details
                val_control = 'Oui' if validation.physical_control_done else 'Non'
                val_notes = validation.control_notes
                val_agent = validation.agent.get_full_name() if validation.agent else ''
                val_agent_code = validation.agent.agent_code if validation.agent else ''
                val_agent_email = validation.agent.email if validation.agent else ''
                val_poe = validation.point_of_exit.name if validation.point_of_exit else ''
                val_poe_code = validation.point_of_exit.code if validation.point_of_exit else ''
                val_poe_type = validation.point_of_exit.get_type_display() if validation.point_of_exit else ''
                val_poe_city = validation.point_of_exit.city if validation.point_of_exit else ''
                val_offline = 'Oui' if validation.is_offline else 'Non'
                val_date = validation.decided_at.strftime('%Y-%m-%d %H:%M') if validation.decided_at else ''
            except:
                val_decision = val_refusal = val_refusal_details = val_control = val_notes = ''
                val_agent = val_agent_code = val_agent_email = ''
                val_poe = val_poe_code = val_poe_type = val_poe_city = val_offline = val_date = ''
            
            # Refund
            try:
                refund = form.refund
                ref_status = refund.get_status_display()
                ref_method = refund.get_method_display()
                ref_gross = refund.gross_amount
                ref_fee = refund.operator_fee
                ref_net = refund.net_amount
                ref_payout_currency = refund.payout_currency or refund.currency
                ref_exchange_rate = float(refund.exchange_rate_applied) if refund.exchange_rate_applied else ''
                ref_payout_amount = float(refund.payout_amount) if refund.payout_amount else ''
                ref_actual_payout = float(refund.actual_payout_amount) if refund.actual_payout_amount else ref_payout_amount
                ref_service_gain = float(refund.service_gain) if refund.service_gain else 0
                ref_service_gain_cdf = float(refund.service_gain_cdf) if refund.service_gain_cdf else 0
                ref_initiated = refund.initiated_at.strftime('%Y-%m-%d %H:%M') if refund.initiated_at else ''
                ref_initiated_by = refund.initiated_by.get_full_name() if refund.initiated_by else ''
                ref_paid = refund.paid_at.strftime('%Y-%m-%d %H:%M') if refund.paid_at else ''
                ref_cash = 'Oui' if refund.cash_collected else 'Non'
                ref_cash_at = refund.cash_collected_at.strftime('%Y-%m-%d %H:%M') if refund.cash_collected_at else ''
                ref_cash_by = refund.cash_collected_by.get_full_name() if refund.cash_collected_by else ''
                ref_attempts = refund.retry_count
            except:
                ref_status = ref_method = ''
                ref_gross = ref_fee = ref_net = ''
                ref_payout_currency = ref_exchange_rate = ref_payout_amount = ref_actual_payout = ''
                ref_service_gain = ref_service_gain_cdf = 0
                ref_initiated = ref_initiated_by = ref_paid = ''
                ref_cash = ref_cash_at = ref_cash_by = ''
                ref_attempts = 0
            
            row_data = [
                form.form_number,
                form.get_status_display(),
                form.currency,
                float(form.eligible_amount),
                float(form.vat_amount),
                float(form.operator_fee),
                float(form.refund_amount),
                form.risk_score,
                'Oui' if form.requires_control else 'Non',
                ', '.join(form.risk_flags) if form.risk_flags else '',
                'Oui' if (form.expires_at and form.expires_at < timezone.now()) else 'Non',
                form.created_at.strftime('%Y-%m-%d %H:%M') if form.created_at else '',
                form.issued_at.strftime('%Y-%m-%d %H:%M') if form.issued_at else '',
                form.expires_at.strftime('%Y-%m-%d %H:%M') if form.expires_at else '',
                form.validated_at.strftime('%Y-%m-%d %H:%M') if form.validated_at else '',
                form.cancelled_at.strftime('%Y-%m-%d %H:%M') if form.cancelled_at else '',
                form.cancellation_reason,
                form.created_by.get_full_name() if form.created_by else '',
                form.cancelled_by.get_full_name() if form.cancelled_by else '',
                # Voyageur
                traveler.last_name,
                traveler.first_name,
                traveler.passport_number_full or f"***{traveler.passport_number_last4}",
                traveler.passport_country,
                traveler.date_of_birth.strftime('%Y-%m-%d') if traveler.date_of_birth else '',
                traveler.nationality,
                traveler.residence_country,
                traveler.email,
                traveler.phone,
                traveler.preferred_refund_method,
                # Facture
                invoice.invoice_number if invoice else '',
                invoice.invoice_date.strftime('%Y-%m-%d') if invoice and invoice.invoice_date else '',
                float(invoice.subtotal) if invoice else '',
                float(invoice.total_vat) if invoice else '',
                float(invoice.total_amount) if invoice else '',
                invoice.payment_method if invoice else '',
                invoice.payment_reference if invoice else '',
                # Commerçant
                merchant.name if merchant else '',
                merchant.trade_name if merchant else '',
                merchant.registration_number if merchant else '',
                merchant.tax_id if merchant else '',
                merchant.city if merchant else '',
                merchant.province if merchant else '',
                merchant.contact_name if merchant else '',
                merchant.contact_email if merchant else '',
                merchant.contact_phone if merchant else '',
                # Point de vente
                outlet.name if outlet else '',
                outlet.code if outlet else '',
                outlet.address_line1 if outlet else '',
                outlet.city if outlet else '',
                # Validation douanière
                val_decision, val_refusal, val_refusal_details, val_control, val_notes,
                val_agent, val_agent_code, val_agent_email,
                val_poe, val_poe_code, val_poe_type, val_poe_city, val_offline, val_date,
                # Remboursement
                ref_status, ref_method, ref_gross, ref_fee, ref_net,
                ref_payout_currency, ref_exchange_rate, ref_payout_amount,
                ref_actual_payout, ref_service_gain, ref_service_gain_cdf,
                ref_initiated, ref_initiated_by, ref_paid,
                ref_cash, ref_cash_at, ref_cash_by, ref_attempts
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_forms.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col in range(1, len(forms_headers) + 1):
            ws_forms.column_dimensions[get_column_letter(col)].width = 15
        
        # ==================== SHEET 2: PRODUITS ====================
        ws_items = wb.create_sheet('Produits')
        
        items_headers = [
            'N° Bordereau', 'N° Facture', 'Commerçant', 'Code Produit', 'Code-barres',
            'Nom Produit', 'Catégorie', 'Description', 'Quantité', 'Prix Unitaire',
            'Total Ligne', 'Taux TVA (%)', 'Montant TVA', 'Éligible', 'Raison Non-éligibilité'
        ]
        
        for col, header in enumerate(items_headers, 1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws_items.freeze_panes = 'A2'
        
        item_row = 2
        for form in queryset:
            if form.invoice:
                for item in form.invoice.items.all():
                    item_data = [
                        form.form_number,
                        form.invoice.invoice_number,
                        form.invoice.merchant.name if form.invoice.merchant else '',
                        item.product_code,
                        item.barcode,
                        item.product_name,
                        item.product_category,
                        item.description,
                        float(item.quantity),
                        float(item.unit_price),
                        float(item.line_total),
                        float(item.vat_rate),
                        float(item.vat_amount),
                        'Oui' if item.is_eligible else 'Non',
                        item.ineligibility_reason
                    ]
                    for col, value in enumerate(item_data, 1):
                        cell = ws_items.cell(row=item_row, column=col, value=value)
                        cell.border = thin_border
                    item_row += 1
        
        for col in range(1, len(items_headers) + 1):
            ws_items.column_dimensions[get_column_letter(col)].width = 15
        
        # ==================== SHEET 3: RÉSUMÉ ====================
        ws_summary = wb.create_sheet('Résumé')
        
        # Calculate summary stats
        total_forms = queryset.count()
        total_eligible = sum(float(f.eligible_amount) for f in queryset)
        total_vat = sum(float(f.vat_amount) for f in queryset)
        total_refund = sum(float(f.refund_amount) for f in queryset)
        total_operator_fee = sum(float(f.operator_fee) for f in queryset)
        
        # Calculate service gains from refunds
        total_service_gain = 0
        total_service_gain_cdf = 0
        for form in queryset:
            try:
                refund = form.refund
                if refund.service_gain:
                    total_service_gain += float(refund.service_gain)
                if refund.service_gain_cdf:
                    total_service_gain_cdf += float(refund.service_gain_cdf)
            except:
                pass
        
        # Get period (date range)
        dates = [f.created_at for f in queryset if f.created_at]
        date_from = min(dates).strftime('%Y-%m-%d') if dates else '-'
        date_to = max(dates).strftime('%Y-%m-%d') if dates else '-'
        
        status_counts = {}
        for form in queryset:
            status = form.get_status_display()
            status_counts[status] = status_counts.get(status, 0) + 1
        
        summary_data = [
            ['RAPPORT D\'EXPORT - BORDEREAUX TAX FREE', ''],
            ['', ''],
            ['Date d\'export', timezone.now().strftime('%Y-%m-%d %H:%M')],
            ['Période', f"Du {date_from} au {date_to}"],
            ['Nombre total de bordereaux', total_forms],
            ['', ''],
            ['MONTANTS TOTAUX', ''],
            ['Montant éligible total', f"{total_eligible:,.2f} CDF"],
            ['TVA totale', f"{total_vat:,.2f} CDF"],
            ['Remboursements totaux', f"{total_refund:,.2f} CDF"],
            ['', ''],
            ['FRAIS ET GAINS', ''],
            ['Total frais opérateur', f"{total_operator_fee:,.2f} CDF"],
            ['Total gain de service', f"{total_service_gain:,.2f}"],
            ['Total gain de service (CDF)', f"{total_service_gain_cdf:,.2f} CDF"],
            ['', ''],
            ['RÉPARTITION PAR STATUT', ''],
        ]
        
        for status, count in status_counts.items():
            summary_data.append([status, count])
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif value in ['MONTANTS TOTAUX', 'FRAIS ET GAINS', 'RÉPARTITION PAR STATUT']:
                    cell.font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bordereaux_export_complet.xlsx"'
        
        return response
    
    def _export_csv(self, request):
        """Fallback CSV export if openpyxl not available."""
        import csv
        from django.http import HttpResponse
        
        queryset = self.get_queryset()[:1000]
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="bordereaux_export.csv"'
        response.write('\ufeff')  # BOM for Excel UTF-8
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'N° Bordereau', 'Statut', 'Voyageur', 'Passeport',
            'Commerçant', 'Point de vente', 'Montant éligible',
            'TVA', 'Remboursement', 'Score risque',
            'Validation', 'Remboursement statut',
            'Créé le', 'Émis le', 'Validé le'
        ])
        
        for form in queryset:
            validation_status = ''
            try:
                validation_status = form.customs_validation.get_decision_display()
            except:
                pass
            
            refund_status = ''
            try:
                refund_status = form.refund.get_status_display()
            except:
                pass
            
            writer.writerow([
                form.form_number,
                form.get_status_display(),
                form.traveler.full_name,
                f"***{form.traveler.passport_number_last4}",
                form.invoice.merchant.name if form.invoice else '',
                form.invoice.outlet.name if form.invoice and form.invoice.outlet else '',
                form.eligible_amount,
                form.vat_amount,
                form.refund_amount,
                form.risk_score,
                validation_status,
                refund_status,
                form.created_at.strftime('%Y-%m-%d %H:%M'),
                form.issued_at.strftime('%Y-%m-%d %H:%M') if form.issued_at else '',
                form.validated_at.strftime('%Y-%m-%d %H:%M') if form.validated_at else '',
            ])
        
        return response


class AdminRefundViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Admin ViewSet for Refunds with full traceability.
    """
    
    queryset = Refund.objects.all()
    permission_classes = [IsAuthenticated, IsAdminOrAuditor]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'method']
    search_fields = [
        'form__form_number',
        'form__traveler__first_name',
        'form__traveler__last_name',
        'form__invoice__merchant__name',
    ]
    ordering_fields = ['created_at', 'initiated_at', 'paid_at', 'net_amount']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AdminRefundDetailSerializer
        return AdminRefundListSerializer

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'form',
            'form__traveler',
            'form__invoice',
            'form__invoice__merchant',
            'form__invoice__outlet',
            'initiated_by',
            'cash_collected_by',
            'cancelled_by',
        ).prefetch_related(
            'attempts',
        )
        
        params = self.request.query_params
        
        # Date range filters
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        # Merchant filter
        merchant_id = params.get('merchant')
        if merchant_id:
            queryset = queryset.filter(form__invoice__merchant_id=merchant_id)
        
        # Outlet filter
        outlet_id = params.get('outlet')
        if outlet_id:
            queryset = queryset.filter(form__invoice__outlet_id=outlet_id)
        
        # Amount range
        min_amount = params.get('min_amount')
        max_amount = params.get('max_amount')
        if min_amount:
            queryset = queryset.filter(net_amount__gte=min_amount)
        if max_amount:
            queryset = queryset.filter(net_amount__lte=max_amount)
        
        # Has retries
        has_retries = params.get('has_retries')
        if has_retries == 'true':
            queryset = queryset.filter(retry_count__gt=0)
        
        return queryset

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get statistics for the refunds dashboard."""
        queryset = self.get_queryset()
        
        params = request.query_params
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        # Status counts
        status_counts = {}
        for status_choice in RefundStatus.choices:
            count = queryset.filter(status=status_choice[0]).count()
            amount = queryset.filter(status=status_choice[0]).aggregate(
                total=Sum('net_amount')
            )['total'] or 0
            status_counts[status_choice[0]] = {
                'count': count,
                'amount': float(amount),
            }
        
        # Method counts
        from apps.refunds.models import RefundMethod
        method_counts = {}
        for method_choice in RefundMethod.choices:
            count = queryset.filter(method=method_choice[0]).count()
            method_counts[method_choice[0]] = count
        
        # Totals
        totals = queryset.aggregate(
            total_count=Count('id'),
            total_gross=Sum('gross_amount'),
            total_fees=Sum('operator_fee'),
            total_net=Sum('net_amount'),
        )
        
        # Paid totals
        paid_totals = queryset.filter(status=RefundStatus.PAID).aggregate(
            count=Count('id'),
            amount=Sum('net_amount'),
        )
        
        # Failed with retries available
        can_retry = queryset.filter(
            status=RefundStatus.FAILED,
            retry_count__lt=3  # max_retries default
        ).count()
        
        return Response({
            'status_counts': status_counts,
            'method_counts': method_counts,
            'totals': {
                'count': totals['total_count'] or 0,
                'gross_amount': float(totals['total_gross'] or 0),
                'fees': float(totals['total_fees'] or 0),
                'net_amount': float(totals['total_net'] or 0),
            },
            'paid': {
                'count': paid_totals['count'] or 0,
                'amount': float(paid_totals['amount'] or 0),
            },
            'can_retry': can_retry,
        })

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export refunds to Excel - Rapport officiel pour services fiscaux."""
        # Check granular permission for EXPORT action
        if request.user.role in ['ADMIN', 'AUDITOR'] and not getattr(request.user, 'is_super_admin', False):
            if not PermissionService.has_module_permission(request.user, 'REFUNDS', 'EXPORT'):
                return Response(
                    {'error': 'Vous n\'avez pas la permission d\'exporter les remboursements.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        import io
        from django.http import HttpResponse
        from decimal import Decimal
        from collections import defaultdict
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return self._export_csv_fallback(request)
        
        # Récupérer les paramètres de période
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        queryset = self.get_queryset().select_related(
            'form__traveler',
            'form__invoice__merchant',
            'form__invoice__outlet',
            'form__customs_validation__agent',
            'form__customs_validation__point_of_exit',
            'initiated_by',
            'cash_collected_by',
            'cancelled_by',
        ).prefetch_related(
            'form__invoice__items',
            'attempts',
        )
        
        # Filtrer par période si spécifié
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        queryset = list(queryset[:2000])
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        subheader_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        title_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        
        money_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
        fee_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        total_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        def style_header(ws, row=1):
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        
        def auto_width(ws):
            for column_cells in ws.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(50, max(12, length + 2))
        
        # Récupérer les paramètres système dynamiques
        from services.settings_service import SettingsService
        system_settings = SettingsService.get_section('general')
        company_name = system_settings.get('company_name', 'Tax Free RDC')
        company_address = system_settings.get('company_address', 'Kinshasa, République Démocratique du Congo')
        company_email = system_settings.get('company_email', 'contact@taxfree.cd')
        company_phone = system_settings.get('company_phone', '+243 XXX XXX XXX')
        
        def add_report_header(ws, title, subtitle=''):
            """Ajouter un en-tête professionnel au rapport"""
            # Titre principal
            ws['A1'] = 'RÉPUBLIQUE DÉMOCRATIQUE DU CONGO'
            ws['A1'].font = Font(bold=True, size=14, color='1E3A5F')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:H1')
            
            ws['A2'] = f'{company_name.upper()} - REMBOURSEMENT TVA AUX VOYAGEURS'
            ws['A2'].font = Font(bold=True, size=12, color='1E3A5F')
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:H2')
            
            ws['A3'] = title
            ws['A3'].font = Font(bold=True, size=11)
            ws['A3'].fill = title_fill
            ws['A3'].font = Font(bold=True, size=11, color='FFFFFF')
            ws['A3'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:H3')
            
            # Informations du rapport
            ws['A5'] = 'Date et heure de génération:'
            ws['B5'] = timezone.now().strftime('%d/%m/%Y à %H:%M:%S')
            ws['B5'].font = Font(bold=True)
            
            # Période
            first_date = queryset[0].created_at.strftime('%d/%m/%Y') if queryset else 'N/A'
            last_date = queryset[-1].created_at.strftime('%d/%m/%Y') if queryset else 'N/A'
            if date_from and date_to:
                period_text = f"Du {date_from} au {date_to}"
            elif queryset:
                period_text = f"Du {first_date} au {last_date}"
            else:
                period_text = "Aucune donnée"
            
            ws['A6'] = 'Période couverte:'
            ws['B6'] = period_text
            ws['B6'].font = Font(bold=True)
            
            ws['A7'] = 'Nombre de transactions:'
            ws['B7'] = len(queryset)
            ws['B7'].font = Font(bold=True)
            
            return 9  # Ligne de départ pour les données
        
        # Calculer les totaux par devise de PAIEMENT (pas devise originale)
        totals_by_currency = defaultdict(lambda: {
            'gross': Decimal('0'), 'fees': Decimal('0'), 'net': Decimal('0'), 'count': 0
        })
        total_gross = Decimal('0')
        total_fees = Decimal('0')
        total_net = Decimal('0')
        paid_count = 0
        conversions_count = 0
        
        for refund in queryset:
            # Utiliser la devise de paiement effective (payout_currency) pour le regroupement
            payout_currency = refund.payout_currency or refund.currency
            
            # Calculer le montant effectivement payé dans la devise de paiement
            if refund.payout_currency and refund.payout_currency != refund.currency:
                # Conversion effectuée - utiliser le montant payé converti
                payout_amount = refund.actual_payout_amount or refund.payout_amount or refund.net_amount
                conversions_count += 1
            else:
                # Pas de conversion - utiliser le montant net
                payout_amount = refund.net_amount
            
            totals_by_currency[payout_currency]['gross'] += refund.gross_amount
            totals_by_currency[payout_currency]['fees'] += refund.operator_fee
            totals_by_currency[payout_currency]['net'] += payout_amount
            totals_by_currency[payout_currency]['count'] += 1
            
            total_gross += refund.gross_amount
            total_fees += refund.operator_fee
            total_net += refund.net_amount
            if refund.status == 'PAID':
                paid_count += 1
        
        # ========== FEUILLE 1: VUE GLOBALE DES REMBOURSEMENTS ==========
        ws1 = wb.active
        ws1.title = 'Vue Globale'
        
        start_row = add_report_header(ws1, 'RAPPORT GLOBAL DES REMBOURSEMENTS')
        
        # En-têtes du tableau principal (sans gain de service - confidentiel)
        headers1 = [
            'N° Bordereau', 'Statut', 'Méthode',
            'Devise', 'Montant Brut (TVA)', 'Frais Opérateur', 'Montant Net',
            'Devise Paiement', 'Taux Change',
            'Voyageur', 'Nationalité',
            'Commerçant', 'Point de Vente',
            'Agent Douane', 'Point de Sortie',
            'Date Paiement'
        ]
        
        for col_idx, header in enumerate(headers1, 1):
            cell = ws1.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        data_start_row = start_row + 1
        
        for refund in queryset:
            form = refund.form
            traveler = form.traveler
            invoice = form.invoice
            merchant = invoice.merchant if invoice else None
            outlet = invoice.outlet if invoice else None
            
            # Customs validation info
            validation_agent = ''
            point_of_exit = ''
            try:
                cv = form.customs_validation
                validation_agent = cv.agent.get_full_name() if cv.agent else ''
                point_of_exit = cv.point_of_exit.name if cv.point_of_exit else ''
            except:
                pass
            
            payout_currency = refund.payout_currency or refund.currency
            exchange_rate = float(refund.exchange_rate_applied) if refund.exchange_rate_applied else '-'
            
            row_data = [
                form.form_number,
                refund.get_status_display(),
                refund.get_method_display(),
                refund.currency,
                float(refund.gross_amount),
                float(refund.operator_fee),
                float(refund.net_amount),
                payout_currency,
                exchange_rate,
                traveler.full_name,
                traveler.nationality,
                merchant.name if merchant else '',
                outlet.name if outlet else '',
                validation_agent,
                point_of_exit,
                refund.paid_at.strftime('%d/%m/%Y %H:%M') if refund.paid_at else '-',
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=data_start_row, column=col_idx, value=value)
                cell.border = thin_border
            
            data_start_row += 1
        
        # Ligne de totaux
        total_row = data_start_row
        ws1.cell(row=total_row, column=1, value='TOTAUX').font = Font(bold=True)
        ws1.cell(row=total_row, column=1).fill = total_fill
        
        for col in range(2, len(headers1) + 1):
            ws1.cell(row=total_row, column=col).fill = total_fill
        
        ws1.cell(row=total_row, column=5, value=float(total_gross)).font = Font(bold=True)
        ws1.cell(row=total_row, column=6, value=float(total_fees)).font = Font(bold=True)
        ws1.cell(row=total_row, column=7, value=float(total_net)).font = Font(bold=True)
        
        # Totaux par devise
        total_row += 2
        ws1.cell(row=total_row, column=1, value='RÉCAPITULATIF PAR DEVISE').font = Font(bold=True, size=11)
        ws1.cell(row=total_row, column=1).fill = subheader_fill
        ws1.cell(row=total_row, column=1).font = Font(bold=True, color='FFFFFF')
        ws1.merge_cells(f'A{total_row}:G{total_row}')
        
        total_row += 1
        for currency, data in totals_by_currency.items():
            ws1.cell(row=total_row, column=1, value=f'Devise: {currency}')
            ws1.cell(row=total_row, column=2, value=f"{data['count']} transactions")
            ws1.cell(row=total_row, column=3, value=f"Brut: {data['gross']:,.2f}")
            ws1.cell(row=total_row, column=4, value=f"Frais: {data['fees']:,.2f}")
            total_row += 1
        
        auto_width(ws1)
        ws1.freeze_panes = f'A{start_row + 1}'
        
        # ========== FEUILLE 2: DÉTAIL DES BORDEREAUX ==========
        ws2 = wb.create_sheet('Détail Bordereaux')
        
        # En-tête
        ws2['A1'] = 'DÉTAIL DES BORDEREAUX TAX FREE'
        ws2['A1'].font = Font(bold=True, size=14, color='1E3A5F')
        ws2.merge_cells('A1:N1')
        
        ws2['A2'] = f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws2['A2'].font = Font(italic=True, color='6B7280')
        
        headers2 = [
            'N° Bordereau', 'Date Création', 'Statut',
            'Voyageur', 'Passeport', 'Nationalité', 'Destination',
            'Commerçant', 'N° RCCM', 'N° Impôt',
            'Point de Vente', 'Ville',
            'Montant Éligible', 'TVA', 'Montant Remboursable',
            'Frais Opérateur', 'Montant Net',
            'Devise Paiement', 'Taux Change',
            'Agent Validation', 'Point de Sortie', 'Date Validation',
            'Méthode Paiement', 'Date Paiement'
        ]
        
        for col_idx, header in enumerate(headers2, 1):
            cell = ws2.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row_num = 5
        total_eligible = Decimal('0')
        total_vat = Decimal('0')
        total_refundable = Decimal('0')
        total_fees_detail = Decimal('0')
        total_net_detail = Decimal('0')
        
        for refund in queryset:
            form = refund.form
            traveler = form.traveler
            invoice = form.invoice
            merchant = invoice.merchant if invoice else None
            outlet = invoice.outlet if invoice else None
            
            # Customs validation info
            validation_agent = ''
            point_of_exit = ''
            validation_date = ''
            try:
                cv = form.customs_validation
                validation_agent = cv.agent.get_full_name() if cv.agent else ''
                point_of_exit = cv.point_of_exit.name if cv.point_of_exit else ''
                validation_date = cv.decided_at.strftime('%d/%m/%Y %H:%M') if cv.decided_at else ''
            except:
                pass
            
            payout_currency = refund.payout_currency or refund.currency
            exchange_rate = float(refund.exchange_rate_applied) if refund.exchange_rate_applied else '-'
            
            total_eligible += form.eligible_amount
            total_vat += form.vat_amount
            total_refundable += form.refund_amount
            total_fees_detail += refund.operator_fee
            total_net_detail += refund.net_amount
            
            row_data = [
                form.form_number,
                form.created_at.strftime('%d/%m/%Y'),
                form.get_status_display(),
                traveler.full_name,
                f"***{traveler.passport_number_last4}",
                traveler.nationality,
                traveler.residence_country or '',
                merchant.name if merchant else '',
                merchant.registration_number if merchant else '',
                merchant.tax_id if merchant else '',
                outlet.name if outlet else '',
                outlet.city if outlet else '',
                float(form.eligible_amount),
                float(form.vat_amount),
                float(form.refund_amount),
                float(refund.operator_fee),
                float(refund.net_amount),
                payout_currency,
                exchange_rate,
                validation_agent,
                point_of_exit,
                validation_date,
                refund.get_method_display(),
                refund.paid_at.strftime('%d/%m/%Y %H:%M') if refund.paid_at else '-',
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col_idx, value=value)
                cell.border = thin_border
            
            row_num += 1
        
        # Ligne de totaux
        ws2.cell(row=row_num, column=1, value='TOTAUX').font = Font(bold=True)
        ws2.cell(row=row_num, column=1).fill = total_fill
        for col in range(2, len(headers2) + 1):
            ws2.cell(row=row_num, column=col).fill = total_fill
        
        ws2.cell(row=row_num, column=13, value=float(total_eligible)).font = Font(bold=True)
        ws2.cell(row=row_num, column=14, value=float(total_vat)).font = Font(bold=True)
        ws2.cell(row=row_num, column=15, value=float(total_refundable)).font = Font(bold=True)
        ws2.cell(row=row_num, column=16, value=float(total_fees_detail)).font = Font(bold=True)
        ws2.cell(row=row_num, column=17, value=float(total_net_detail)).font = Font(bold=True)
        
        auto_width(ws2)
        ws2.freeze_panes = 'A5'
        
        # ========== FEUILLE 3: PRODUITS VENDUS ==========
        ws3 = wb.create_sheet('Produits Vendus')
        
        ws3['A1'] = 'DÉTAIL DES PRODUITS VENDUS'
        ws3['A1'].font = Font(bold=True, size=14, color='1E3A5F')
        ws3.merge_cells('A1:N1')
        
        headers3 = [
            'N° Bordereau', 'Date', 
            'Commerçant', 'N° RCCM', 'N° NIF', 'N° Id. Nat.',
            'Point de Vente',
            'Produit', 'Code-barres', 'Catégorie',
            'Quantité', 'Prix Unitaire', 'Total HT',
            'Taux TVA (%)', 'Montant TVA', 'Total TTC',
            'Éligible au remboursement'
        ]
        
        for col_idx, header in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row_num = 4
        total_ht = Decimal('0')
        total_tva_products = Decimal('0')
        total_ttc = Decimal('0')
        
        for refund in queryset:
            form = refund.form
            invoice = form.invoice
            if not invoice:
                continue
            
            merchant = invoice.merchant
            merchant_name = merchant.name if merchant else ''
            merchant_rccm = merchant.registration_number if merchant else ''
            merchant_tax_id = merchant.tax_id if merchant else ''
            # Récupérer national_id depuis la demande d'inscription originale
            merchant_national_id = ''
            if merchant:
                from apps.accounts.models import MerchantRegistrationRequest
                reg_request = MerchantRegistrationRequest.objects.filter(
                    registration_number=merchant.registration_number
                ).first()
                if reg_request:
                    merchant_national_id = reg_request.national_id or ''
            outlet_name = invoice.outlet.name if invoice.outlet else ''
            
            for item in invoice.items.all():
                item_ttc = float(item.line_total) + float(item.vat_amount)
                total_ht += item.line_total
                total_tva_products += item.vat_amount
                total_ttc += Decimal(str(item_ttc))
                
                row_data = [
                    form.form_number,
                    form.created_at.strftime('%d/%m/%Y'),
                    merchant_name,
                    merchant_rccm,
                    merchant_tax_id,
                    merchant_national_id,
                    outlet_name,
                    item.product_name,
                    item.barcode or '',
                    item.product_category or '',
                    float(item.quantity),
                    float(item.unit_price),
                    float(item.line_total),
                    float(item.vat_rate),
                    float(item.vat_amount),
                    item_ttc,
                    'Oui' if item.is_eligible else 'Non',
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws3.cell(row=row_num, column=col_idx, value=value)
                    cell.border = thin_border
                
                row_num += 1
        
        # Totaux produits
        ws3.cell(row=row_num, column=1, value='TOTAUX').font = Font(bold=True)
        ws3.cell(row=row_num, column=1).fill = total_fill
        for col in range(2, len(headers3) + 1):
            ws3.cell(row=row_num, column=col).fill = total_fill
        
        ws3.cell(row=row_num, column=10, value=float(total_ht)).font = Font(bold=True)
        ws3.cell(row=row_num, column=12, value=float(total_tva_products)).font = Font(bold=True)
        ws3.cell(row=row_num, column=13, value=float(total_ttc)).font = Font(bold=True)
        
        auto_width(ws3)
        ws3.freeze_panes = 'A4'
        
        # Note: Feuille 4 (Tentatives) supprimée - non pertinente pour rapport fiscal
        
        # ========== FEUILLE 4: RÉSUMÉ FISCAL (sans gains - confidentiel) ==========
        ws4 = wb.create_sheet('Résumé Fiscal')
        
        # En-tête officiel
        ws4['A1'] = 'RÉPUBLIQUE DÉMOCRATIQUE DU CONGO'
        ws4['A1'].font = Font(bold=True, size=14, color='1E3A5F')
        ws4['A1'].alignment = Alignment(horizontal='center')
        ws4.merge_cells('A1:E1')
        
        ws4['A2'] = f'{company_name.upper()} - RÉCAPITULATIF FISCAL'
        ws4['A2'].font = Font(bold=True, size=12, color='1E3A5F')
        ws4['A2'].alignment = Alignment(horizontal='center')
        ws4.merge_cells('A2:E2')
        
        ws4['A4'] = f"Date de génération: {timezone.now().strftime('%d/%m/%Y à %H:%M:%S')}"
        ws4['A4'].font = Font(italic=True, color='6B7280')
        
        # Période
        first_date = queryset[0].created_at.strftime('%d/%m/%Y') if queryset else 'N/A'
        last_date = queryset[-1].created_at.strftime('%d/%m/%Y') if queryset else 'N/A'
        if date_from and date_to:
            period_text = f"Du {date_from} au {date_to}"
        elif queryset:
            period_text = f"Du {first_date} au {last_date}"
        else:
            period_text = "Aucune donnée"
        ws4['A5'] = f"Période couverte: {period_text}"
        ws4['A5'].font = Font(bold=True)
        
        # Stats générales
        ws4['A7'] = 'STATISTIQUES GÉNÉRALES'
        ws4['A7'].font = Font(bold=True, color='FFFFFF')
        ws4['A7'].fill = subheader_fill
        ws4.merge_cells('A7:E7')
        
        row = 8
        ws4[f'A{row}'] = 'Nombre total de remboursements'
        ws4[f'B{row}'] = len(queryset)
        row += 1
        ws4[f'A{row}'] = 'Remboursements effectués (payés)'
        ws4[f'B{row}'] = paid_count
        row += 1
        ws4[f'A{row}'] = 'Taux de succès'
        ws4[f'B{row}'] = f"{(paid_count / len(queryset) * 100):.1f}%" if queryset else '0%'
        row += 1
        ws4[f'A{row}'] = 'Remboursements avec conversion de devise'
        ws4[f'B{row}'] = conversions_count
        row += 2
        
        # Montants totaux
        ws4[f'A{row}'] = 'MONTANTS TOTAUX'
        ws4[f'A{row}'].font = Font(bold=True, color='FFFFFF')
        ws4[f'A{row}'].fill = subheader_fill
        ws4.merge_cells(f'A{row}:E{row}')
        row += 1
        
        ws4[f'A{row}'] = 'Total TVA remboursable (montant brut)'
        ws4[f'B{row}'] = f"{total_gross:,.2f}"
        ws4[f'C{row}'] = 'CDF'
        ws4[f'B{row}'].fill = money_fill
        row += 1
        ws4[f'A{row}'] = 'Total frais opérateur perçus'
        ws4[f'B{row}'] = f"{total_fees:,.2f}"
        ws4[f'C{row}'] = 'CDF'
        ws4[f'B{row}'].fill = fee_fill
        row += 1
        ws4[f'A{row}'] = 'Total net remboursé aux voyageurs'
        ws4[f'B{row}'] = f"{total_net:,.2f}"
        ws4[f'C{row}'] = 'CDF'
        ws4[f'B{row}'].fill = money_fill
        row += 2
        
        # Totaux par devise
        ws4[f'A{row}'] = 'RÉPARTITION PAR DEVISE'
        ws4[f'A{row}'].font = Font(bold=True, color='FFFFFF')
        ws4[f'A{row}'].fill = subheader_fill
        ws4.merge_cells(f'A{row}:E{row}')
        row += 1
        
        ws4[f'A{row}'] = 'Devise Paiement'
        ws4[f'B{row}'] = 'Transactions'
        ws4[f'C{row}'] = 'Montant Brut (CDF)'
        ws4[f'D{row}'] = 'Frais (CDF)'
        for col in ['A', 'B', 'C', 'D']:
            ws4[f'{col}{row}'].font = Font(bold=True)
        row += 1
        
        for currency, data in totals_by_currency.items():
            ws4[f'A{row}'] = currency
            ws4[f'B{row}'] = data['count']
            ws4[f'C{row}'] = f"{data['gross']:,.2f} CDF"
            ws4[f'D{row}'] = f"{data['fees']:,.2f} CDF"
            row += 1
        row += 1
        
        # Par méthode de paiement
        ws4[f'A{row}'] = 'PAR MÉTHODE DE PAIEMENT'
        ws4[f'A{row}'].font = Font(bold=True, color='FFFFFF')
        ws4[f'A{row}'].fill = subheader_fill
        ws4.merge_cells(f'A{row}:E{row}')
        row += 1
        
        from apps.refunds.models import RefundMethod
        for method_choice in RefundMethod.choices:
            count = sum(1 for r in queryset if r.method == method_choice[0])
            amount = sum(r.net_amount for r in queryset if r.method == method_choice[0])
            ws4[f'A{row}'] = str(method_choice[1])
            ws4[f'B{row}'] = f"{count} transactions"
            ws4[f'C{row}'] = f"{amount:,.2f} CDF"
            row += 1
        row += 1
        
        # Par statut
        ws4[f'A{row}'] = 'PAR STATUT DE REMBOURSEMENT'
        ws4[f'A{row}'].font = Font(bold=True, color='FFFFFF')
        ws4[f'A{row}'].fill = subheader_fill
        ws4.merge_cells(f'A{row}:E{row}')
        row += 1
        
        for status_choice in RefundStatus.choices:
            count = sum(1 for r in queryset if r.status == status_choice[0])
            amount = sum(r.net_amount for r in queryset if r.status == status_choice[0])
            ws4[f'A{row}'] = str(status_choice[1])
            ws4[f'B{row}'] = f"{count} transactions"
            ws4[f'C{row}'] = f"{amount:,.2f} CDF"
            row += 1
        row += 1
        
        # Par commerçant (Top 10)
        ws4[f'A{row}'] = 'PAR COMMERÇANT (Top 10)'
        ws4[f'A{row}'].font = Font(bold=True, color='FFFFFF')
        ws4[f'A{row}'].fill = subheader_fill
        ws4.merge_cells(f'A{row}:E{row}')
        row += 1
        
        merchant_stats = {}
        for refund in queryset:
            merchant_name = refund.form.invoice.merchant.name if refund.form.invoice and refund.form.invoice.merchant else 'Inconnu'
            if merchant_name not in merchant_stats:
                merchant_stats[merchant_name] = {'count': 0, 'amount': Decimal('0'), 'fees': Decimal('0')}
            merchant_stats[merchant_name]['count'] += 1
            merchant_stats[merchant_name]['amount'] += refund.net_amount
            merchant_stats[merchant_name]['fees'] += refund.operator_fee
        
        sorted_merchants = sorted(merchant_stats.items(), key=lambda x: x[1]['amount'], reverse=True)[:10]
        for merchant_name, data in sorted_merchants:
            ws4[f'A{row}'] = merchant_name
            ws4[f'B{row}'] = f"{data['count']} transactions"
            ws4[f'C{row}'] = f"{data['amount']:,.2f} CDF"
            ws4[f'D{row}'] = f"Frais: {data['fees']:,.2f} CDF"
            row += 1
        
        auto_width(ws4)
        
        # ========== FEUILLE 5: ATTESTATION LÉGALE ==========
        ws5 = wb.create_sheet('Attestation Légale')
        
        ws5['A1'] = 'RÉPUBLIQUE DÉMOCRATIQUE DU CONGO'
        ws5['A1'].font = Font(bold=True, size=14, color='1E3A5F')
        ws5['A1'].alignment = Alignment(horizontal='center')
        ws5.merge_cells('A1:F1')
        
        ws5['A2'] = 'ATTESTATION DE REMBOURSEMENTS TVA'
        ws5['A2'].font = Font(bold=True, size=12)
        ws5['A2'].alignment = Alignment(horizontal='center')
        ws5.merge_cells('A2:F2')
        
        ws5['A4'] = 'Le présent document certifie les remboursements de TVA effectués dans le cadre'
        ws5['A5'] = f'du système {company_name} conformément à la réglementation en vigueur.'
        
        first_date = queryset[0].created_at.strftime('%d/%m/%Y') if queryset else 'N/A'
        last_date = queryset[-1].created_at.strftime('%d/%m/%Y') if queryset else 'N/A'
        if date_from and date_to:
            period_text = f"Du {date_from} au {date_to}"
        elif queryset:
            period_text = f"Du {first_date} au {last_date}"
        else:
            period_text = "Aucune donnée"
        
        ws5['A7'] = 'Période couverte:'
        ws5['B7'] = period_text
        ws5['B7'].font = Font(bold=True)
        ws5['A8'] = 'Date de génération:'
        ws5['B8'] = timezone.now().strftime('%d/%m/%Y à %H:%M:%S')
        ws5['B8'].font = Font(bold=True)
        
        ws5['A10'] = 'RÉCAPITULATIF FISCAL'
        ws5['A10'].font = Font(bold=True, size=11)
        ws5['A10'].fill = subheader_fill
        ws5['A10'].font = Font(bold=True, color='FFFFFF')
        ws5.merge_cells('A10:C10')
        
        ws5['A12'] = 'Total TVA remboursée:'
        ws5['B12'] = f"{total_gross:,.2f} CDF"
        ws5['B12'].font = Font(bold=True)
        ws5['A13'] = 'Frais de service perçus:'
        ws5['B13'] = f"{total_fees:,.2f} CDF"
        ws5['B13'].font = Font(bold=True)
        ws5['A14'] = 'Montant net versé aux voyageurs:'
        ws5['B14'] = f"{total_net:,.2f} CDF"
        ws5['B14'].font = Font(bold=True)
        ws5['A15'] = 'Nombre de transactions:'
        ws5['B15'] = len(queryset)
        ws5['B15'].font = Font(bold=True)
        
        ws5['A17'] = 'MENTION LÉGALE'
        ws5['A17'].font = Font(bold=True)
        ws5['A18'] = 'Ce document constitue un justificatif officiel des remboursements de TVA effectués'
        ws5['A19'] = f'dans le cadre du système {company_name}. Il peut être présenté aux services fiscaux'
        ws5['A20'] = 'de la République Démocratique du Congo comme preuve des opérations réalisées.'
        
        ws5['A22'] = f'Document généré automatiquement par le système {company_name}.'
        ws5['A23'] = f'Contact: {company_email} | Tél: {company_phone}'
        ws5['A23'].font = Font(italic=True, color='6B7280')
        ws5['A24'] = f'Adresse: {company_address}'
        ws5['A24'].font = Font(italic=True, color='6B7280')
        ws5['A22'].font = Font(italic=True, color='6B7280')
        
        auto_width(ws5)
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"rapport_remboursements_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def _export_csv_fallback(self, request):
        """Fallback CSV export if openpyxl is not available."""
        import csv
        from django.http import HttpResponse
        
        queryset = self.get_queryset()[:1000]
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="remboursements_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'N° Bordereau', 'Voyageur', 'Commerçant', 'Point de vente',
            'Montant brut', 'Frais', 'Montant net', 'Devise',
            'Méthode', 'Statut', 'Initié par', 'Initié le',
            'Payé le', 'Tentatives', 'Créé le'
        ])
        
        for refund in queryset:
            writer.writerow([
                refund.form.form_number,
                refund.form.traveler.full_name,
                refund.form.invoice.merchant.name if refund.form.invoice else '',
                refund.form.invoice.outlet.name if refund.form.invoice and refund.form.invoice.outlet else '',
                refund.gross_amount,
                refund.operator_fee,
                refund.net_amount,
                refund.currency,
                refund.get_method_display(),
                refund.get_status_display(),
                refund.initiated_by.full_name if refund.initiated_by else '',
                refund.initiated_at.strftime('%Y-%m-%d %H:%M') if refund.initiated_at else '',
                refund.paid_at.strftime('%Y-%m-%d %H:%M') if refund.paid_at else '',
                refund.retry_count,
                refund.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        
        return response
